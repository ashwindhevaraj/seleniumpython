from openpyxl import load_workbook

workbook = load_workbook('Data2.xlsx')
sheet=workbook['Sheet1']
print(sheet.max_column)
print(sheet.max_row)
username=sheet['A2'].value
password=sheet['B2'].value
print(username)
print(password)

for row in sheet.iter_rows(min_row=2, values_only=True):
    username=row[0]
    password=row[1]
    print(username,password)